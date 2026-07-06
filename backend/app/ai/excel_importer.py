from io import BytesIO
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.ai.statement_importer import ImportError


def xlsx_to_text(raw: bytes) -> str:
    """
    Flatten an .xlsx workbook into a comma/newline text block so it can be
    fed through the same text-based extraction prompt CSV uses.

    Raises ImportError if the bytes aren't a readable .xlsx file.
    """
    try:
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, KeyError, ParseError) as exc:
        raise ImportError(
            "Could not read this Excel file. Make sure it's a valid .xlsx export."
        ) from exc

    multi_sheet = len(workbook.sheetnames) > 1
    blocks: list[str] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        lines = [
            ",".join("" if cell is None else str(cell) for cell in row)
            for row in sheet.iter_rows(values_only=True)
        ]
        block = "\n".join(lines)
        if multi_sheet:
            block = f"# Sheet: {sheet_name}\n{block}"
        blocks.append(block)

    return "\n\n".join(blocks)
