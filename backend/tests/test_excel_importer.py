# tests/test_excel_importer.py
#
# Builds small workbooks in memory with openpyxl (no fixture files needed)
# and verifies xlsx_to_text flattens them into the comma/newline text format
# that statement_importer's text prompt expects.

from io import BytesIO
from zipfile import ZipFile

import pytest
from openpyxl import Workbook

from app.ai.excel_importer import xlsx_to_text
from app.ai.statement_importer import ImportError as StatementImportError


def _make_xlsx(sheets: dict[str, list[list]]) -> bytes:
    """sheets: {sheet_name: [[row1_cells], [row2_cells], ...]}"""
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestXlsxToText:
    def test_flattens_rows_into_comma_joined_text(self):
        raw = _make_xlsx({
            "Sheet1": [
                ["Date", "Description", "Amount"],
                ["2024-03-15", "Coffee shop", 4.5],
            ]
        })
        text = xlsx_to_text(raw)
        assert "Date,Description,Amount" in text
        assert "2024-03-15,Coffee shop,4.5" in text

    def test_empty_cells_become_empty_string(self):
        raw = _make_xlsx({"Sheet1": [["A", None, "C"]]})
        text = xlsx_to_text(raw)
        assert "A,,C" in text

    def test_multiple_sheets_are_prefixed_with_sheet_name(self):
        raw = _make_xlsx({
            "Jan": [["a", "b"]],
            "Feb": [["c", "d"]],
        })
        text = xlsx_to_text(raw)
        assert "# Sheet: Jan" in text
        assert "# Sheet: Feb" in text
        assert "a,b" in text
        assert "c,d" in text

    def test_single_sheet_has_no_sheet_prefix(self):
        raw = _make_xlsx({"OnlySheet": [["a", "b"]]})
        text = xlsx_to_text(raw)
        assert "# Sheet:" not in text

    def test_corrupt_bytes_raise_import_error(self):
        with pytest.raises(StatementImportError):
            xlsx_to_text(b"this is not a real xlsx file")

    def test_malformed_xml_inside_valid_zip_raises_import_error(self):
        """
        Test that a valid ZIP container with malformed/truncated internal XML
        (a realistic corruption case) raises ImportError instead of ParseError.
        """
        # Create a valid .xlsx first
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Test"
        ws['B1'] = "Data"
        ws.append([1, 2])
        buf = BytesIO()
        wb.save(buf)
        valid_xlsx_bytes = buf.getvalue()

        # Corrupt the internal XML by replacing workbook.xml with malformed XML
        zip_buffer = BytesIO()
        with ZipFile(BytesIO(valid_xlsx_bytes), 'r') as orig_zip:
            with ZipFile(zip_buffer, 'w') as new_zip:
                for item in orig_zip.infolist():
                    data = orig_zip.read(item.filename)
                    if item.filename == 'xl/workbook.xml':
                        # Truncate with malformed XML that will trigger ParseError in openpyxl
                        data = b"<Workbook><incomplete xml"
                    new_zip.writestr(item, data)

        corrupted_xlsx_bytes = zip_buffer.getvalue()

        # Should raise ImportError, not ParseError
        with pytest.raises(StatementImportError):
            xlsx_to_text(corrupted_xlsx_bytes)
